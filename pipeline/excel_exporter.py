import re
from openpyxl import Workbook

# ── Helper: parse a dimension string like "4ft-6in", "7in", "3ft" → value in feet ──
def _parse_dim_to_feet(dim_str):
    """Convert a dimension string to a numeric value in feet. Returns None on failure."""
    if not dim_str or not isinstance(dim_str, str):
        return None
    dim_str = dim_str.strip().lower().replace("'", "ft").replace('"', 'in').replace("–", "-").replace("—", "-")
    # Try ft + in pattern: "4ft-6in", "4ft 6in", "4ft-0in"
    m = re.match(r'(\d+(?:\.\d+)?)\s*ft[\s\-]*(\d+(?:\.\d+)?)\s*in', dim_str)
    if m:
        return float(m.group(1)) + float(m.group(2)) / 12.0
    # Try ft only: "4ft", "4 ft"
    m = re.match(r'(\d+(?:\.\d+)?)\s*ft', dim_str)
    if m:
        return float(m.group(1))
    # Try in only: "54in", "7 in"
    m = re.match(r'(\d+(?:\.\d+)?)\s*in', dim_str)
    if m:
        return float(m.group(1)) / 12.0
    # Try mm: "150mm"
    m = re.match(r'(\d+(?:\.\d+)?)\s*mm', dim_str)
    if m:
        return float(m.group(1)) / 304.8
    # Try bare number (assume inches)
    m = re.match(r'^(\d+(?:\.\d+)?)$', dim_str)
    if m:
        return float(m.group(1)) / 12.0
    return None


#.......................................
def create_excel_from_result(result: dict, output_path: str):

    wb = Workbook()

    line_items = result.get("line_items", [])
    confidence = result.get("confidence")
    grand_total = result.get("grand_total")
    survey_data = result.get("survey_data", {})
    project_specs = result.get("project_specs", {})

    # -------------------------------
    # Calculate Total Deductions
    # -------------------------------
    total_deductions = sum(
        item.get("Total_SF", 0)
        for item in line_items
        if str(item.get("Category", "")).strip().lower() == "deduction"
    )

    # ===============================
    # SUMMARY SHEET
    # ===============================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Grand Total SF", grand_total])
    ws_summary.append(["Total Deductions SF", total_deductions])
    ws_summary.append(["Confidence", confidence])

    # ===============================
    # DEDUCTIONS SHEET
    # ===============================
    ws_ded = wb.create_sheet("Deductions")

    ws_ded.append([
        "Page",
        "View",
        "Description",
        "Dimensions",
        "EA",
        "Unit_SF",
        "Total_SF",
    ])

    for item in line_items:
        if str(item.get("Category", "")).strip().lower() == "deduction":
            ws_ded.append([
                item.get("Page"),
                item.get("View"),
                item.get("Description"),
                item.get("Dimensions", ""),
                item.get("Count"),
                item.get("Unit_SF"),
                item.get("Total_SF"),
            ])


    # -------------------------------
    # Fallback: Check for missing specs from Survey Data
    # -------------------------------
    # Create a set of processed items (Page, View, clean_tag) to avoid duplicates
    processed_keys = set()
    for item in line_items:
        if str(item.get("Category", "")).strip().lower() == "deduction":
            # Extract tag from description "Opening TYPE A" -> "TYPE A"
            desc = item.get("Description", "")
            # Simple heuristic: assume description ends with the tag or contains it
            processed_keys.add(f"{item.get('Page')}_{item.get('View')}_{desc}")

    for page_num, views in survey_data.items():
        for view_label, tags in views.items():
             for tag, count in tags.items():
                  # Clean tag logic from phase5_v2 reuse
                  clean_tag = str(tag).upper().replace("TYPE", "").strip()
                  
                  # Check if we have a robust match in processed_keys
                  # Since we can't easily reconstruct the exact Description phase5 used, 
                  # we'll look for fuzzy match or just see if "Opening" + Clean Tag is plausible
                  
                  # Actually, easier: check if this tag exists in the specs. 
                  # If NOT, it's definitely missing.
                  
                  # Flatten library (lite version of phase5 logic)
                  all_types = {}
                  all_types.update(project_specs.get('windows', {}))
                  all_types.update(project_specs.get('doors', {}))
                  
                  spec_found = False
                  if clean_tag in all_types:
                      spec_found = True
                  else:
                      for key in all_types:
                          if clean_tag == key.replace("-", "") or key == clean_tag.replace("-", ""):
                              spec_found = True
                              break
                  
                  if not spec_found:
                      ws_ded.append([
                        page_num,
                        view_label,
                        f"Opening {clean_tag}",
                        "MISSING",
                        count,
                        0.0,
                        0.0
                      ])
    # ===============================
    # EIFS SHEET
    # ===============================
    ws_eifs = wb.create_sheet("EIFS")

    ws_eifs.append([
        "Page",
        "View",
        "Description",
        "Dimensions",
        "EA",
        "Unit_SF",
        "Total_SF",
    ])

    for item in line_items:
        category = item.get("Category", "")

        if "EIFS" in category.upper():
            ws_eifs.append([
                item.get("Page"),
                item.get("View"),
                item.get("Description"),
                item.get("Dimensions", ""),
                item.get("Count"),
                item.get("Unit_SF"),
                item.get("Total_SF"),
            ])

    # ── Fascia & Reveal extractions ──────────────────────────────────────────
    for extraction_key, label in [("fascia_extraction", "Fascia"), ("reveal_extraction", "Reveal")]:
        extraction = result.get(extraction_key)
        if not extraction or not isinstance(extraction, dict):
            continue
        for page_entry in extraction.get("page_results", []):
            page_num = page_entry.get("page")
            if page_num is not None:
                page_num = page_num + 1          # 0-based → 1-based
            keyword = page_entry.get("keyword", label)
            res = page_entry.get("result", {})
            if res.get("status") != "SUCCESS":
                continue

            for occ in res.get("occurrence_results", [res]):
                p2 = occ.get("phase2", {}) or {}
                p3 = occ.get("phase3", {}) or {}
                p7 = occ.get("phase7", {}) or {}

                drawing_title = p2.get("drawing_title", "")
                material      = p3.get("material", "")
                height_val    = p7.get("height")
                width_val     = p7.get("width")

                # ── Build dimensions string ──
                # Priority: phase7 height/width → phase3 dimension_label_text
                dim_parts = []
                if height_val:
                    dim_parts.append(str(height_val))
                if width_val:
                    dim_parts.append(str(width_val))
                dimensions = " x ".join(dim_parts) if dim_parts else (
                    p3.get("dimension_label_text") or ""
                )

                # ── Compute Unit_SF ──
                # Priority 1: phase3 numeric height_value + unit
                # Priority 2: parse phase7 height string
                # Priority 3: parse phase3 dimension_label_text
                unit_sf = None
                h = p3.get("height_value")
                unit = p3.get("unit", "")
                if h is not None:
                    try:
                        h_num = float(h)
                        if unit and unit.lower() in ("in", "inch", "inches"):
                            h_num = h_num / 12.0
                        unit_sf = round(h_num, 4)
                    except (ValueError, TypeError):
                        pass

                # Fallback: parse phase7 height string
                if unit_sf is None and height_val:
                    parsed_ft = _parse_dim_to_feet(str(height_val))
                    if parsed_ft is not None:
                        unit_sf = round(parsed_ft, 4)

                # Fallback: parse phase3 dimension_label_text
                if unit_sf is None:
                    dim_text = p3.get("dimension_label_text") or ""
                    if dim_text:
                        parsed_ft = _parse_dim_to_feet(dim_text)
                        if parsed_ft is not None:
                            unit_sf = round(parsed_ft, 4)

                # ── Description ──
                description = f"{keyword}"
                if material:
                    description += f" - {material}"

                ws_eifs.append([
                    page_num,
                    drawing_title or "",
                    description,
                    dimensions,
                    1,
                    unit_sf,
                    unit_sf,        # Total = Unit for single occurrence
                ])

    wb.save(output_path)
